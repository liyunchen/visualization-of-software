# -*- coding: utf-8 -*-


"""
李运辰 2021-3-34

公众号：python爬虫数据分析挖掘
"""

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.3; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0',}



# 1
#7973227714515400
# 2
#4779805474835700
# 3
#/37/00/1016845483273700
# 4
#8679935826337700
# 5
#7197533339804600
# 6
#8042213977424800
# 7
#2262609044749800
# 8
#1699488619659400
# 9
#1805374511564700
# 10
#1933721047194600
# 11
#7232026471250800
# 12
#8982352350925900
# 13
#4702797553454300
# 14
#2151107991923800
# 15
#8357465155589300
# 16
#2071693573022900
# 17
#4646645944127100
# 18
#1182091647913900
# 19
#7711721648193100
# 20
#2099769377685800
# 21
#3042314248738300
# 22
#2889100571832100
# 23
#3374410909698000
# 24
#4335405595243700
# 25
#5215381530163200
# 26
#2379725258541100
# 27
#4872856713204800
# 28
#1488519001760800

tv_name_list =[
'',
'/54/00/7973227714515400',
'/57/00/4779805474835700',
'/37/00/1016845483273700',
'/77/00/8679935826337700',
'/46/00/7197533339804600',
'/48/00/8042213977424800',
'/98/00/2262609044749800',
'/94/00/1699488619659400',
'/47/00/1805374511564700',
'/46/00/1933721047194600',
'/08/00/7232026471250800',
'/59/00/8982352350925900',
'/43/00/4702797553454300',
'/38/00/2151107991923800',
'/93/00/8357465155589300',
'/29/00/2071693573022900',
'/71/00/4646645944127100',
'/39/00/1182091647913900',
'/31/00/7711721648193100',
'/58/00/2099769377685800',
'/83/00/3042314248738300',
'/21/00/2889100571832100',
'/98/00/3374410909698000',
'/37/00/4335405595243700',
'/32/00/5215381530163200',
'/11/00/2379725258541100',
'/48/00/4872856713204800',
'/08/00/1488519001760800',
]

import openpyxl
outwb = openpyxl.Workbook()  # 打开一个将写的文件
outws = outwb.create_sheet(index=0)  # 在将写的文件创建sheet

"""
import xlwt
# # 创建一个workbook 设置编码
workbook = xlwt.Workbook(encoding = 'utf-8')
# # 创建一个worksheet
worksheet = workbook.add_sheet('sheet1')
#
# # 写入excel
# # 参数对应 行, 列, 值
# worksheet.write(0,0, label='index')
# worksheet.write(0,1, label='tvname')
# worksheet.write(0,2, label='uid')
# worksheet.write(0,3, label='content')
# worksheet.write(0,4, label='likeCount')
"""

outws.cell(row = 1 , column = 1 , value = "index")
outws.cell(row = 1 , column = 2 , value = "tvname")
outws.cell(row = 1 , column = 3 , value = "uid")
outws.cell(row = 1 , column = 4 ,  value = "content")
outws.cell(row = 1 , column = 5 , value = "likeCount")


import zlib
import requests


# 1.爬取xml文件
def download_xml(url):
    bulletold = requests.get(url).content  # 二进制内容
    return zipdecode(bulletold)

def zipdecode(bulletold):
    '对zip压缩的二进制内容解码成文本'
    decode = zlib.decompress(bytearray(bulletold), 15 + 32).decode('utf-8')
    return decode

def get_data():
    for k in range(1,len(tv_name_list)):#29个 1-28
        url_id = tv_name_list[k]
        for x in range(1,11):
            # x是从1到11，11怎么来的，这一集总共46分钟，爱奇艺每5分钟会加载新的弹幕,46除以5向上取整
            try:
                url = 'https://cmts.iqiyi.com/bullet'+str(url_id)+'_300_' + str(x) + '.z'
                xml = download_xml(url)
                # 把编码好的文件分别写入个xml文件中（类似于txt文件），方便后边取数据
                with open('./lyc/zx'+str(k) +'-'+ str(x) + '.xml', 'a+', encoding='utf-8') as f:
                    f.write(xml)
            except:
                pass


count = 2
# 2.读取xml文件中的弹幕数据数据
from xml.dom.minidom import parse
import xml.dom.minidom
def xml_parse(file_name,tv__name):
    global  count
    DOMTree = xml.dom.minidom.parse(file_name)
    collection = DOMTree.documentElement
    # 在集合中获取所有entry数据
    entrys = collection.getElementsByTagName("entry")

    for entry in entrys:
        uid = entry.getElementsByTagName('uid')[0]
        content = entry.getElementsByTagName('content')[0]
        likeCount = entry.getElementsByTagName('likeCount')[0]
        #print(uid.childNodes[0].data)
        #print(content.childNodes[0].data)
        #print(likeCount.childNodes[0].data)
        # 写入excel
        # 参数对应 行, 列, 值
        outws.cell(row=count, column=1, value=str(count))
        outws.cell(row=count, column=2, value=str("第"+str(tv__name)+"集"))
        outws.cell(row=count, column=3, value=str(uid.childNodes[0].data))
        outws.cell(row=count, column=4, value=str(content.childNodes[0].data))
        outws.cell(row=count, column=5, value=str(likeCount.childNodes[0].data))
        count=count+1


def combine_data():
    for k in range(1,29):
        for x in range(1,11):
            try:
                xml_parse("./lyc/zx"+str(k) +"-"+ str(x) + ".xml",k)
                print(str(k) + "-" + str(x))
            except:
                pass
    # 保存
    #workbook.save('弹幕数据集-李运辰.xls')
    outwb.save("弹幕数据集-李运辰.xls")  # 保存结果


###获取数据
#get_data()

###合并写excel
combine_data()